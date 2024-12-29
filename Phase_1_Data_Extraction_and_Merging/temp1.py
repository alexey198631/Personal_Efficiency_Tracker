import gspread
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from data_files import config
from datetime import datetime

filename = config.SERVICE_ACCOUNT
dairy_file = config.DAYS
gbook = config.GBOOK
glist = config.GSHEET
elist = config.ELIST
gtiming = config.TIMINGBOOK

# Prepare data frame for Excel workbook sheet
def df_from_sheet(wkb, sht_name):
    # Select the sheet you want to modify
    sheet = wkb[sht_name]

    # Save data frame from sheet data
    data_days = sheet.values
    cols = next(data_days)[0:]
    df = pd.DataFrame(data_days, columns=cols)
    return df


# Write the updated DataFrame to the worksheet with the same format
def same_format_sheet(wkb, sht, df):
    # Create a new worksheet with the same formatting as the existing worksheet
    new_worksheet = wkb.create_sheet('Temp')
    new_worksheet.sheet_format = sht.sheet_format
    new_worksheet.sheet_properties = sht.sheet_properties
    new_worksheet.page_setup = sht.page_setup

    # Write the updated DataFrame to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        new_worksheet.append(r)

    # Delete all existing rows in the worksheet
    sht.delete_rows(1, sht.max_row)
    for row in new_worksheet.iter_rows():
        sht.append([cell.value for cell in row])

    # Delete the temporary worksheet
    wkb.remove(new_worksheet)

    # Save the changes to the Excel file
    wkb.save(dairy_file)

    # Close the workbook
    wkb.close()


# Open a Google sheet
gc = gspread.service_account(filename)
spr = gc.open(gbook).worksheet(gtiming)


gday = config.DAYBOOK
gweek = config.WEEKBOOK
gmonth = config.MONTHBOOK

def sheet_to_df(sheet_var):
    # Open a Google sheet
    spr = gc.open(gbook).worksheet(sheet_var)
    # Get all the data from the sheet as a list of lists
    data = spr.get_all_values()
    # Convert the data to a pandas dataframe
    df_google = pd.DataFrame(data[1:], columns=data[0])
    return df_google

df = sheet_to_df(gtiming)
#days = sheet_to_df(gday)
#weeks = sheet_to_df(gweek)
#months = sheet_to_df(gmonth)

# Define an integer format string
number_format = '#,##0.00'

#writer = pd.ExcelWriter('data_files/timing.xlsx', engine='xlsxwriter')
#df.to_excel(writer, sheet_name='Timing', index=False, float_format=number_format)
#days.to_excel(writer, sheet_name='Days', index=False, float_format=number_format)
#weeks.to_excel(writer, sheet_name='Weeks', index=False, float_format=number_format)
#months.to_excel(writer, sheet_name='Months', index=False, float_format=number_format)
#writer.close()


def change_column_types(df, column_types, time_format='%H:%M'):
    """
    Change the types of columns in a pandas DataFrame.

    Parameters:
    - df: pandas DataFrame.
    - column_types: Dictionary with column names as keys and the target type as values.
                    Acceptable types are 'integer', 'float','time', 'text', 'categorical', 'date'.
    - time_format: String specifying the time format, default is '%H:%M:%S'.

    Returns:
    - DataFrame with changed column types.
    """
    for column, target_type in column_types.items():
        if target_type == 'integer':
            df[column] = pd.to_numeric(df[column], errors='coerce').astype('float')
        elif target_type == 'float':
            df[column] = pd.to_numeric(df[column], errors='coerce').astype('float')
        elif target_type == 'time':
            df[column] = pd.to_datetime(df[column], format=time_format).dt.time
        elif target_type == 'date':
            pass
            #df[column] = pd.to_datetime(df[column], format='%Y-%m-%d').dt.time
        elif target_type == 'text':
            df[column] = df[column].astype(str)
        elif target_type == 'categorical':
            df[column] = pd.Categorical(df[column])
        else:
            raise ValueError(f"Unsupported type: {target_type}")
    return df

# Identify the first occurrence of an empty value in the 'Event' column
first_empty_index = df[df['EVENT'] == ''].index[0]

# Slice the DataFrame up to the row before the first empty value
df = df.iloc[:first_empty_index]

# Formatting
df_cleaned = df[['DATE', 'START', 'FINISH', 'DURATION', 'TYPE', 'DIRECTION', 'EVENT',
       'LANGUAGE', 'LANG TYPE', 'REMARK', 'IN/OUT', 'P/Y', 'WALK', 'RUN',
       'CYCLE', 'PUSH_UPS', 'PULL_UPS', 'SKID', 'SQUATING', 'ABS', 'PLANK',
       'WATER', 'GOALS', 'DUMBBELLS', 'FASTFOOD', 'SWEETS', 'BEER', 'WINE',
       'COCTAIL', 'VODKA', 'WHISKY', 'BRANDY', 'COFFEE', 'ART', 'YEAR', 'SIZE',
       'CREATOR', 'WHERE', 'COUNTRY', 'PTS', 'COMMENTS', 'TYPE.1']]


df_cleaned[['WALK', 'RUN',
       'CYCLE', 'PUSH_UPS', 'PULL_UPS', 'SKID', 'SQUATING', 'ABS', 'PLANK',
       'WATER', 'GOALS', 'DUMBBELLS', 'FASTFOOD', 'SWEETS', 'BEER', 'WINE',
       'COCTAIL', 'VODKA', 'WHISKY', 'BRANDY', 'COFFEE', 'ART', 'YEAR', 'SIZE',
       'CREATOR', 'WHERE', 'COUNTRY', 'PTS', 'COMMENTS', 'TYPE.1']] = df_cleaned[['WALK', 'RUN',
       'CYCLE', 'PUSH_UPS', 'PULL_UPS', 'SKID', 'SQUATING', 'ABS', 'PLANK',
       'WATER', 'GOALS', 'DUMBBELLS', 'FASTFOOD', 'SWEETS', 'BEER', 'WINE',
       'COCTAIL', 'VODKA', 'WHISKY', 'BRANDY', 'COFFEE', 'ART', 'YEAR', 'SIZE',
       'CREATOR', 'WHERE', 'COUNTRY', 'PTS', 'COMMENTS', 'TYPE.1']].fillna(0)


column_types = {
    'DATE':'date',
    'START':'time',
    'FINISH':'time',
    'DURATION':'integer',
    'TYPE':'text',
    'DIRECTION':'text',
    'EVENT':'text',
    'LANGUAGE':'text',
    'LANG TYPE':'text',
    'REMARK':'text',
    'IN/OUT':'text',
    'P/Y':'text',
    'WALK':'text',
    'RUN':'text',
    'CYCLE':'text',
    'PUSH_UPS':'integer',
    'PULL_UPS':'integer',
    'SKID':'integer',
    'SQUATING':'integer',
    'ABS':'integer',
    'PLANK':'integer',
    'WATER':'integer',
    'GOALS':'float',
    'DUMBBELLS':'integer',
    'FASTFOOD':'integer',
    'SWEETS':'integer',
    'BEER':'integer',
    'WINE':'integer',
    'COCTAIL':'integer',
    'VODKA':'integer',
    'WHISKY':'integer',
    'BRANDY':'integer',
    'COFFEE':'integer',
    'ART':'text',
    'YEAR':'integer',
    'SIZE':'integer',
    'CREATOR':'text',
    'WHERE':'text',
    'COUNTRY':'text',
    'PTS':'float',
    'COMMENTS':'text',
    'TYPE.1':'text'
}


# Drop rows where the first column (column 'A' in this example) is NaN.
df_cleaned = df_cleaned.dropna(subset=['DURATION'])
df_cleaned = change_column_types(df_cleaned, column_types)


# Replace empty strings with NaN in the 'Event' column
#df_cleaned.loc[:, 'WALK'] = df_cleaned['WALK'].replace('', np.nan)
#df_cleaned.loc[:, 'RUN'] = df_cleaned['RUN'].replace('', np.nan)
#df_cleaned.loc[:, 'CYCLE'] = df_cleaned['CYCLE'].replace('', np.nan)

# Replace empty strings with NaN and convert to float
df_cleaned.loc[:, 'WALK'] = pd.to_numeric(df_cleaned['WALK'].replace('', np.nan), errors='coerce')
df_cleaned.loc[:, 'RUN'] = pd.to_numeric(df_cleaned['RUN'].replace('', np.nan), errors='coerce')
df_cleaned.loc[:, 'CYCLE'] = pd.to_numeric(df_cleaned['CYCLE'].replace('', np.nan), errors='coerce')


df_cleaned[['PUSH_UPS', 'PULL_UPS', 'SKID', 'SQUATING', 'ABS', 'PLANK',
       'WATER', 'GOALS', 'DUMBBELLS', 'FASTFOOD', 'SWEETS', 'BEER', 'WINE',
       'COCTAIL', 'VODKA', 'WHISKY', 'BRANDY', 'COFFEE', 'ART', 'YEAR', 'SIZE',
       'CREATOR', 'WHERE', 'COUNTRY', 'PTS', 'COMMENTS', 'TYPE.1']] = df_cleaned[['PUSH_UPS', 'PULL_UPS', 'SKID', 'SQUATING', 'ABS', 'PLANK',
       'WATER', 'GOALS', 'DUMBBELLS', 'FASTFOOD', 'SWEETS', 'BEER', 'WINE',
       'COCTAIL', 'VODKA', 'WHISKY', 'BRANDY', 'COFFEE', 'ART', 'YEAR', 'SIZE',
       'CREATOR', 'WHERE', 'COUNTRY', 'PTS', 'COMMENTS', 'TYPE.1']].fillna(0)


# Art sheet preparation
df_ARTS = df[['DATE', 'START', 'FINISH', 'DURATION', 'TYPE', 'DIRECTION', 'EVENT',
       'LANGUAGE', 'LANG TYPE', 'REMARK', 'IN/OUT', 'P/Y', 'WALK', 'RUN',
       'CYCLE', 'PUSH_UPS', 'PULL_UPS', 'SKID', 'SQUATING', 'ABS', 'PLANK',
       'WATER', 'GOALS', 'DUMBBELLS', 'FASTFOOD', 'SWEETS', 'BEER', 'WINE',
       'COCTAIL', 'VODKA', 'WHISKY', 'BRANDY', 'COFFEE', 'ART', 'YEAR', 'SIZE',
       'CREATOR', 'WHERE', 'COUNTRY', 'PTS', 'COMMENTS', 'TYPE.1']]

# Convert the 'Event' column to string type to handle potential new categories
df_ARTS['DURATION'] = df_ARTS['DURATION'].astype(float)

df_for_art = df_ARTS.copy()
df_for_art = df_for_art.drop('TYPE', axis=1)
df_for_art = df_for_art.rename(columns={'ART':'NAME','TYPE.1': 'TYPE', 'DURATION': 'TIME_SPENT', 'LANGUAGE': 'LANG','LANG TYPE': 'ACTIVITY'})
df_for_art['AWARD'] = 0
df_for_art['TIMES'] = 1

# Replace empty strings with NaN in the 'Event' column
df_for_art.loc[:, 'NAME'] = df_for_art['NAME'].replace('', np.nan)
df_for_art.loc[:, 'TYPE'] = df_for_art['TYPE'].replace('', np.nan)

# Drop rows where the first column (column 'ART', 'YEAR' in this example) is NaN.
df_art = df_for_art.dropna(subset=['NAME'])
df_for_art = df_art[['DATE', 'LANG','TYPE', 'AWARD', 'TIMES', 'NAME', 'YEAR', 'SIZE', 'ACTIVITY','TIME_SPENT','CREATOR', 'WHERE', 'COUNTRY', 'PTS', 'COMMENTS']]
# Creating the pivot table
art_pivot_df = df_for_art.pivot_table(index=['NAME'], values='TIME_SPENT', aggfunc='sum').reset_index()

# Drop rows where the first column (column 'ART', 'YEAR' in this example) is NaN.
df_art = df_for_art.dropna(subset=['NAME', 'TYPE'])

arts = df_art['NAME'].to_list()

# Drop rows where the first column (column 'A' in this example) is NaN.
tempdf = df_cleaned.copy()
tempdf.loc[:, 'ART'] = tempdf['ART'].replace('', np.nan)

df_art_started = tempdf.dropna(subset=['ART'])
df_art_started = df_art_started.drop('TYPE', axis=1)

# Keeping rows where values in columna are not in the exclusion_list
df_filtered = df_art_started[~df_art_started['ART'].isin(arts)]

df_filtered_finished = df_art_started[df_art_started['ART'].isin(arts)]

# Keeping rows where values in columna are not in the exclusion_list
df_filtered_no_duplicates = df_filtered.drop_duplicates(subset=['ART'], keep='first')
df_filtered_no_duplicates_finished = df_filtered_finished.drop_duplicates(subset=['ART'], keep='first')
df_filtered_no_duplicates_finished = df_filtered_no_duplicates_finished[['DATE', 'ART']]
df_filtered_no_duplicates_finished = df_filtered_no_duplicates_finished.rename(columns={'DATE':'START_DATE', 'ART':'NAME'})

df_filtered_no_duplicates = df_filtered_no_duplicates.rename(columns={'ART':'NAME','TYPE.1': 'TYPE', 'DURATION': 'TIME_SPENT', 'LANGUAGE': 'LANG','LANG TYPE': 'ACTIVITY'})
df_filtered_no_duplicates['AWARD'] = 0
df_filtered_no_duplicates['TIMES'] = 1

df_for_art_started = df_filtered_no_duplicates[['DATE', 'LANG', 'TYPE', 'AWARD', 'TIMES', 'NAME', 'YEAR', 'SIZE', 'ACTIVITY','CREATOR', 'WHERE', 'COUNTRY', 'PTS', 'COMMENTS']]

df_for_art_started = pd.merge(df_for_art_started, art_pivot_df, how='left', on='NAME')
df_for_art_started['END_DATE'] = pd.Timestamp(datetime.now())
df_for_art_started = df_for_art_started.rename(columns={'DATE':'START_DATE'})

# Convert the start_date column to datetime format
df_for_art_started['START_DATE'] = pd.to_datetime(df_for_art_started['START_DATE'], format='%d.%m.%Y')

df_for_art_started['DAYS'] = (datetime.now() - df_for_art_started['START_DATE']).dt.days + 1


df_for_art_started = df_for_art_started[['START_DATE', 'END_DATE','DAYS', 'LANG','TYPE', 'AWARD', 'TIMES', 'NAME', 'YEAR', 'SIZE', 'ACTIVITY', 'TIME_SPENT','CREATOR', 'WHERE', 'COUNTRY', 'PTS', 'COMMENTS']]

list_df_started_not_finished = df_for_art_started['NAME'].to_list()

df_art = df_art.drop('TIME_SPENT', axis=1)


df_for_art_finished = pd.merge(df_art, art_pivot_df, how='left', on='NAME')
df_for_art_finished = df_for_art_finished.rename(columns={'DATE':'END_DATE'})
df_for_art_finished = pd.merge(df_for_art_finished, df_filtered_no_duplicates_finished, on='NAME', how='left')

# Convert the 'START_DATE' and 'END_DATE' columns to datetime using the specified format
df_for_art_finished['START_DATE'] = pd.to_datetime(df_for_art_finished['START_DATE'], format='%d.%m.%Y')
df_for_art_finished['END_DATE'] = pd.to_datetime(df_for_art_finished['END_DATE'], format='%d.%m.%Y')

df_for_art_finished['DAYS'] = (df_for_art_finished['END_DATE'] - df_for_art_finished['START_DATE']).dt.days + 1
df_for_art_finished = df_for_art_finished[['START_DATE', 'END_DATE','DAYS','LANG','TYPE', 'AWARD', 'TIMES', 'NAME', 'YEAR', 'SIZE', 'ACTIVITY', 'TIME_SPENT','CREATOR', 'WHERE', 'COUNTRY', 'PTS', 'COMMENTS']]

# Calculate the number of full days between the dates
art_final = pd.concat([df_for_art_finished, df_for_art_started])

art_final['SIZE'] = art_final['SIZE'].astype(int)
# Replace commas with dots in the 'VALUE' column and convert to float
art_final['PTS'] = art_final['PTS'].str.replace(',', '.').astype(float)

# error with ,
df_cleaned['RUN'] = df_cleaned['RUN'].astype(str)
df_cleaned['CYCLE'] = df_cleaned['CYCLE'].astype(str)
df_cleaned['WALK'] = df_cleaned['WALK'].astype(str)

df_cleaned['RUN'] = df_cleaned['RUN'].str.replace(',', '.').astype(float)
df_cleaned['CYCLE'] = df_cleaned['CYCLE'].str.replace(',', '.').astype(float)
df_cleaned['WALK'] = df_cleaned['WALK'].str.replace(',', '.').astype(float)

df_cleaned[['WALK', 'RUN',
       'CYCLE']] = df_cleaned[['WALK', 'RUN',
       'CYCLE']].fillna(0)


# PYTHON

df_python = df_cleaned[df_cleaned['EVENT'].str.contains('Python', case=False, na=False)]
df_python = df_python[['DATE', 'DURATION']]
# Creating the pivot table
df_python = df_python.pivot_table(index='DATE', values='DURATION', aggfunc='sum', fill_value=0)
df_python = df_python.reset_index()
df_python = df_python.rename(columns={'DURATION': 'PYTHON'})

# LANGUAGES
df_language = df_cleaned.dropna(subset=['LANGUAGE'])

# Creating the pivot table
language_table = df_language.pivot_table(
    index='DATE',
    columns=['LANG TYPE', 'LANGUAGE'],
    values='DURATION',
    aggfunc='sum',
    fill_value=0
)

# Flatten the columns
language_table.columns = ['{}_{}'.format(type_, lang) for type_, lang in language_table.columns]
# Reset the index to turn the date index into a column (optional, depending on your needs)
language_table.reset_index(inplace=True)

# Extract unique languages from column names
unique_languages = set(lang.split('_')[-1] for lang in language_table.columns if '_' in lang)

# Sum durations for each language across types and add as new TOTAL columns
for lang in unique_languages:
    # Construct column names for this language across all types
    lang_columns = [col for col in language_table.columns if col.endswith(f'_{lang}')]
    # Sum these columns and create a new TOTAL column for the language
    language_table[f'TOTAL_{lang}'] = language_table[lang_columns].sum(axis=1)

# clean WRITing

clean_writing_df = df_cleaned.dropna(subset=['LANGUAGE'])
clean_writing_df = clean_writing_df[clean_writing_df['LANG TYPE'] == 'WRITE']
clean_writing_df = clean_writing_df[clean_writing_df['TYPE'] != 'meets']
# Remove unused categories
clean_writing_df['LANG TYPE'] = clean_writing_df['LANG TYPE'].astype('category')

clean_writing_df['LANG TYPE'] = clean_writing_df['LANG TYPE'].cat.remove_unused_categories()
clean_writing_pivot = clean_writing_df.pivot_table(
    index='DATE',
    columns=['LANG TYPE'],
    values='DURATION',
    aggfunc='sum',
    fill_value=0,
    observed=True
)

clean_writing_pivot.reset_index(inplace=True)

# Rename the column 'OldColumnName' to 'NewColumnName'
clean_writing_pivot.rename(columns={'WRITE': 'WRITE_J'}, inplace=True)

# Merge the DataFrames on 'DATE' using an outer join to include all dates
language_table = pd.merge(language_table, clean_writing_pivot, on='DATE', how='left')

# VALUES COLUMNS

columns_values = ['WALK', 'RUN', 'CYCLE', 'PUSH_UPS', 'PULL_UPS', 'SKID', 'SQUATING', 'ABS', 'PLANK',
                  'WATER', 'GOALS', 'DUMBBELLS', 'FASTFOOD', 'SWEETS', 'BEER', 'WINE',
                  'COCTAIL', 'VODKA', 'WHISKY', 'BRANDY', 'COFFEE']

# and select only the 'DATE', 'DURATION', and 'EVENT' columns.
df_sleeping = df_cleaned[df_cleaned['TYPE'] == 'sleeping'][['DATE', 'START', 'FINISH', 'DURATION', 'EVENT']]

# and select only the 'DATE', 'DURATION', and 'EVENT' columns.
df_morning_place = df_cleaned[df_cleaned['TYPE'] == 'morning'][['DATE', 'EVENT']]
df_morning_place = df_morning_place.rename(columns={'EVENT': 'PL1'})
df_sport_place = df_cleaned[df_cleaned['TYPE'] == 'sport'][['DATE', 'EVENT']]
df_sport_place = df_sport_place.rename(columns={'EVENT': 'PL2'})
df_places = pd.merge(df_morning_place, df_sport_place, on='DATE', how='outer')


# Define the function to apply conditions for the 'REGIME' column
def calculate_regime(row):
    regime = 0
    start = row['FINISH']
    finish = row['START']

    # Define start and end times for comparison
    start_time_lower = pd.to_datetime('05:00:00', format='%H:%M:%S').time()
    start_time_upper = pd.to_datetime('07:15:00', format='%H:%M:%S').time()
    finish_time_lower = pd.to_datetime('20:00:00', format='%H:%M:%S').time()
    finish_time_upper = pd.to_datetime('23:15:00', format='%H:%M:%S').time()

    # Check if 'START' is between 5:00 AM and 7:15 AM
    if start_time_lower <= start <= start_time_upper:
        regime += 0.5

    # Check if 'FINISH' is between 20:00 PM and 23:15 PM
    if finish_time_lower <= finish <= finish_time_upper:
        regime += 0.5

    return regime


# Calculate 'hours' as 'DURATION' divided by 8 and round to one decimal place
df_sleeping['hours'] = (df_sleeping['DURATION']).round(0)
# Convert 'START' and 'FINISH' columns to datetime to easily work with time
df_sleeping['START'] = pd.to_datetime(df_sleeping['START'], format='%H:%M:%S').dt.time
df_sleeping['FINISH'] = pd.to_datetime(df_sleeping['FINISH'], format='%H:%M:%S').dt.time
# Apply the function to each row to calculate the 'REGIME' value
df_sleeping['REGIME'] = df_sleeping.apply(calculate_regime, axis=1)

# and select only the 'DATE', 'DURATION', and 'EVENT' columns.
df_visualization = df_cleaned[df_cleaned['REMARK'] == 'V'][['DATE']]
df_visualization['Visalization'] = 1

# and select only the 'DATE', 'DURATION', and 'EVENT' columns.
df_pnw = df_cleaned[df_cleaned['REMARK'] == 'P'][['DATE']]
df_pnw['PNW'] = 0

# and select only the 'DATE', 'DURATION', and 'EVENT' columns.
df_alw = df_cleaned[df_cleaned['REMARK'] == 'A'][['DATE']]
df_alw['ALW'] = 0

# and select only the 'DATE', 'DURATION', and 'EVENT' columns.
df_outside = df_cleaned[df_cleaned['IN/OUT'] == 'OUT'][['DATE', 'DURATION']]
# Creating the pivot table
outside_pivot_df = df_outside.pivot_table(index='DATE', values='DURATION', aggfunc='sum', fill_value=0)

outside_pivot_df = outside_pivot_df.reset_index()
outside_pivot_df = outside_pivot_df.rename(columns={'DURATION': 'OUTSIDE'})

# Merge the DataFrames on 'DATE' using an outer join to include all dates
df_merged = pd.merge(df_sleeping, outside_pivot_df, on='DATE', how='left')

# and select only the 'DATE', 'DURATION', and 'EVENT' columns.
df_eyes = df_cleaned[df_cleaned['TYPE'] == 'eyes'][['DATE', 'DURATION']]
# Creating the pivot table
df_eyes = df_eyes.pivot_table(index='DATE', values='DURATION', aggfunc='sum', fill_value=0)

df_eyes = df_eyes.reset_index()
df_eyes = df_eyes.rename(columns={'DURATION': 'EYES_MIN'})

# Merge the DataFrames on 'DATE' using an outer join to include all dates
df_merged = pd.merge(df_merged, df_eyes, on='DATE', how='left')

# Merge Language data
df_merged = pd.merge(df_merged, language_table, on='DATE', how='left')
# Merge Python data
df_merged = pd.merge(df_merged, df_python, on='DATE', how='left')

# and select only the 'DATE', 'DURATION', and 'EVENT' columns.
df_paid_work = df_cleaned[df_cleaned['P/Y'] == 'Y'][['DATE', 'DURATION']]
df_paid_work = df_paid_work.pivot_table(index='DATE', values='DURATION', aggfunc='sum', fill_value=0)
df_paid_work = df_paid_work.reset_index()
df_paid_work = df_paid_work.rename(columns={'DURATION': 'PAID_WORK'})

# Merge the DataFrames on 'DATE' using an outer join to include all dates
df_merged = pd.merge(df_merged, df_paid_work, on='DATE', how='left')

# Merge the DataFrames on 'DATE' using an outer join to include all dates
df_merged = pd.merge(df_merged, df_places, on='DATE', how='left')

columns_pivots_df = df_cleaned.pivot_table(index='DATE', values=columns_values, aggfunc='sum', fill_value=0)
columns_pivots_df = columns_pivots_df.reset_index()

# List of columns you want to make negative
columns_to_negate = ['FASTFOOD', 'SWEETS', 'BEER', 'WINE',
                     'COCTAIL', 'VODKA', 'WHISKY', 'BRANDY', 'COFFEE']

# Make the selected columns negative
columns_pivots_df[columns_to_negate] = columns_pivots_df[columns_to_negate] * -1

# Merge the DataFrames on 'DATE' using an outer join to include all dates
df_merged = pd.merge(df_merged, columns_pivots_df, on='DATE', how='left')

# Merge the DataFrames on 'DATE' using an outer join to include all dates
df_merged = pd.merge(df_merged, df_visualization, on='DATE', how='left')

# Fill missing values in column 'V' with 0
df_merged['Visalization'] = df_merged['Visalization'].fillna(0)

# Merge the DataFrames on 'DATE' using an outer join to include all dates
df_merged = pd.merge(df_merged, df_pnw, on='DATE', how='left')

# Fill missing values in column 'P' with 1
df_merged['PNW'] = df_merged['PNW'].fillna(1)

# Merge the DataFrames on 'DATE' using an outer join to include all dates
df_merged = pd.merge(df_merged, df_alw, on='DATE', how='left')

# Fill missing values in column 'P' with 1
df_merged['ALW'] = df_merged['ALW'].fillna(1)

# List of 'TYPE' values you're interested in
type_values_sum = ['sleeping', 'break', 'breath', 'morning', 'operations', 'meets',
                   'days', 'language', 'prof', 'meditation', 'feats', 'prior', 'iq', 'plans', 'sport']

type_values_count = ['morning', 'eyes', 'qo', 'sport', 'ql', 'ctrain']

# Creating the pivot table
pivot_df = df_cleaned.pivot_table(index='DATE', columns='TYPE', values='DURATION', aggfunc='sum', fill_value=0)

# Filter the pivot table to only include the specified 'TYPE' columns
# This step ensures all your desired 'TYPE' values are included even if they're not present in the data
pivot_df = pivot_df.reindex(columns=type_values_sum, fill_value=0)

pivot_df = pivot_df.reset_index()

# Merge the DataFrames on 'DATE' using an outer join to include all dates
df_merged = pd.merge(df_merged, pivot_df, on='DATE', how='left')

# paid types
clean_working_df = df_cleaned[df_cleaned['P/Y'] == 'Y']

# Change values in column 'TYPE' from 'days' to 'plans'
clean_working_df.loc[clean_working_df['TYPE'] == 'days', 'TYPE'] = 'plans'
clean_working_df.loc[clean_working_df['TYPE'] == 'eyes', 'TYPE'] = 'prof'
clean_working_df.loc[clean_working_df['TYPE'] == 'breath', 'TYPE'] = 'prof'
clean_working_df.loc[clean_working_df['TYPE'] == 'iq', 'TYPE'] = 'meets'
clean_working_df.loc[clean_working_df['TYPE'] == 'language', 'TYPE'] = 'meets'
# Creating the pivot table
pivot_clean_working_df = clean_working_df.pivot_table(index='DATE', columns='TYPE', values='DURATION', aggfunc='sum',
                                                      fill_value=0)
work_type_values_sum = ['operations', 'meets', 'prof', 'feats', 'prior', 'plans']
pivot_clean_working_df = pivot_clean_working_df.reindex(columns=work_type_values_sum, fill_value=0)
pivot_clean_working_df = pivot_clean_working_df.reset_index()

# Convert column names to uppercase and add prefix "Y_"
pivot_clean_working_df.rename(columns=lambda x: 'Y_' + x.upper(), inplace=True)

# Rename the column
pivot_clean_working_df.rename(columns={'Y_DATE': 'DATE'}, inplace=True)

# Merge the DataFrames on 'DATE' using an outer join to include all dates
df_merged = pd.merge(df_merged, pivot_clean_working_df, on='DATE', how='left')

df_merged = df_merged.drop_duplicates(subset=['DATE'])

df_merged['SPORT_TIME'] = df_merged['morning'] + df_merged['sport']
df_merged['PLANS'] = df_merged['plans'] + df_merged['days']

# Creating a pivot table to count occurrences of each 'TYPE' for each 'DATE'
pivot_df = df_cleaned.pivot_table(index='DATE', columns='TYPE', aggfunc='size', fill_value=0)

# Ensure all desired 'TYPE' columns are included, even if they're not present in the original data
pivot_df = pivot_df.reindex(columns=type_values_count, fill_value=0)

columns = ['morning']

# Apply the condition to each specified column
for col in columns:
    pivot_df[col] = pivot_df[col].apply(lambda x: 1 if x >= 1 else 0)

# Merge the DataFrames on 'DATE' using an outer join to include all dates
df_merged = pd.merge(df_merged, pivot_df, on='DATE', how='left')
df_merged = df_merged.drop_duplicates(subset=['DATE'])

df_merged['PL1'] = df_merged['PL1'].fillna(0)
df_merged['PL2'] = df_merged['PL2'].fillna(0)
df_merged['WRITE_J'] = df_merged['WRITE_J'].fillna(0)
df_merged['EYES_MIN'] = df_merged['EYES_MIN'].fillna(0)

# Fill NaN values with 0 in specific columns
columns_to_fill = ['WRITE_J', 'Y_OPERATIONS', 'Y_MEETS', 'Y_PROF', 'Y_FEATS', 'Y_PRIOR', 'Y_PLANS']
df_merged[columns_to_fill] = df_merged[columns_to_fill].fillna(0)

df_merged['PAID_WORK'] = df_merged['PAID_WORK'].fillna(0)
df_merged['PYTHON'] = df_merged['PYTHON'].fillna(0)

df_merged = df_merged[
    ['DATE', 'START', 'FINISH', 'REGIME', 'hours', 'Visalization', 'meditation', 'breath', 'ctrain', 'morning_y',
     'sport_y', 'SPORT_TIME', 'PL1', 'PL2', 'OUTSIDE', 'eyes', 'EYES_MIN', 'ql', 'qo', 'PNW', 'ALW', 'WALK', 'RUN',
     'CYCLE', 'PUSH_UPS', 'PULL_UPS', 'SKID', 'SQUATING', 'ABS', 'PLANK',
     'WATER', 'GOALS', 'DUMBBELLS', 'FASTFOOD', 'SWEETS', 'BEER', 'WINE',
     'COCTAIL', 'VODKA', 'WHISKY', 'BRANDY', 'COFFEE', 'break', 'iq', 'prof', 'language', 'prior', 'PLANS',
     'operations', 'meets', 'feats', 'morning_x', 'sport_x', 'LIST_DUT', 'LIST_ENG', 'LIST_RUS', 'LIST_SPA', 'READ_DUT',
     'READ_ENG', 'READ_RUS', 'READ_SPA', 'SPEAK_DUT', 'SPEAK_ENG',
     'SPEAK_RUS', 'SPEAK_SPA', 'TEST_DUT', 'TEST_ENG', 'TEST_RUS',
     'TEST_SPA', 'WATCH_DUT', 'WATCH_ENG', 'WATCH_RUS', 'WATCH_SPA',
     'WRITE_DUT', 'WRITE_ENG', 'WRITE_RUS', 'WRITE_SPA', 'TOTAL_SPA',
     'TOTAL_DUT', 'TOTAL_RUS', 'TOTAL_ENG', 'WRITE_J', 'PYTHON', 'PAID_WORK', 'Y_OPERATIONS', 'Y_MEETS', 'Y_PROF',
     'Y_FEATS', 'Y_PRIOR',
     'Y_PLANS']]


# Renaming the column 'oldName1' to 'newName1'
df_final = df_merged.rename(
    columns={'hours': 'SLEEP', 'meditation': 'MEDITATION', 'breath': 'BREATH', 'ctrain': 'CTRAINING',
             'morning_y': 'MORNING', 'sport_y': 'SPORT', 'eyes': 'EYES', 'ql': 'MK', 'qo': 'PT', 'iq': 'IQ',
             'prof': 'PROF', 'prior': 'PRIOR', 'operations': 'ROUTINE', 'meets': 'MEETS', 'feats': 'DIFFICULT',
             'morning_x': 'MORNING_TIME', 'sport_x': 'DAY_TRAINING_TIME', 'language': 'LANGUAGE', 'break': 'BREAK'})

# additional calculations
df_final['MK_MIN'] = df_final['MK'] * 11
df_final['PT_MIN'] = df_final['PT'] * 7


yes_no = ['REGIME', 'Visalization', 'CTRAINING', 'MORNING', 'PNW', 'ALW']
times = ['SPORT', 'EYES', 'MK', 'PT']
negative = ['FASTFOOD', 'SWEETS', 'COFFEE']
quantities = ['PUSH_UPS', 'PULL_UPS', 'SKID', 'SQUATING', 'ABS', 'PLANK', 'GOALS', 'DUMBBELLS']
distance = ['WALK', 'RUN', 'CYCLE', 'WATER']
separate_timing = ['SLEEP', 'BREAK','MEDITATION', 'BREATH', 'MK_MIN', 'PT_MIN', 'EYES_MIN', 'SPORT_TIME', 'IQ', 'PROF', 'LANGUAGE', 'PRIOR', 'PLANS', 'ROUTINE', 'MEETS', 'DIFFICULT']
imposed_timing = ['OUTSIDE', 'TOTAL_ENG', 'TOTAL_RUS', 'TOTAL_SPA', 'TOTAL_DUT', 'PAID_WORK']
additional = ['PYTHON', 'WRITE_J', 'Y_PRIOR', 'Y_PLANS','Y_OPERATIONS', 'Y_MEETS', 'Y_PROF', 'Y_FEATS']
typ_lang_eng = ['LIST_ENG', 'READ_ENG', 'SPEAK_ENG', 'WATCH_ENG', 'TEST_ENG', 'WRITE_ENG']
typ_lang_rus = ['LIST_RUS', 'READ_RUS', 'SPEAK_RUS', 'WATCH_RUS', 'TEST_ENG', 'WRITE_RUS']
typ_lang_spa = ['LIST_SPA', 'READ_SPA', 'SPEAK_SPA', 'WATCH_SPA', 'TEST_SPA', 'WRITE_SPA']
typ_lang_dut = ['LIST_DUT', 'READ_DUT', 'SPEAK_DUT', 'WATCH_DUT', 'TEST_DUT', 'WRITE_DUT']
sport_additional = ['MORNING_TIME','DAY_TRAINING_TIME', 'STRENGHT_POINTS']

listening = ['LIST_ENG', 'LIST_RUS', 'LIST_SPA', 'LIST_DUT']
reading = ['READ_ENG', 'READ_RUS', 'READ_SPA', 'READ_DUT']
watching = ['WATCH_ENG', 'WATCH_RUS', 'WATCH_SPA', 'WATCH_DUT']
testing = ['TEST_DUT', 'TEST_ENG', 'TEST_SPA'] #'TEST_RUS'
speaking = ['SPEAK_ENG', 'SPEAK_RUS', 'SPEAK_SPA', 'SPEAK_DUT']
writing = ['WRITE_ENG', 'WRITE_RUS', 'WRITE_SPA', 'WRITE_DUT']


df_final['TOTAL_TIME'] = df_final[separate_timing].sum(axis=1)
separate_timing = ['TOTAL_TIME', 'SLEEP', 'BREAK', 'MEDITATION', 'BREATH', 'MK_MIN', 'PT_MIN', 'EYES_MIN', 'SPORT_TIME', 'IQ', 'PROF', 'LANGUAGE', 'PRIOR', 'PLANS', 'ROUTINE', 'MEETS', 'DIFFICULT']


df_final['STRENGHT_POINTS'] = df_final.PUSH_UPS + df_final.PULL_UPS * 3 + df_final.SKID * 2 + df_final.SQUATING + df_final.ABS + df_final.PLANK / 2 + df_final.WATER / 4 + df_final.RUN * 50 + df_final.CYCLE * 20 + df_final.WALK * 50 + df_final.GOALS * 10 + df_final.DUMBBELLS / 2

# Create a new column 'SUM' which is the sum of the specified columns
df_final['SUM_LIST'] = df_final[listening].sum(axis=1)
df_final['SUM_READ'] = df_final[reading].sum(axis=1)
df_final['SUM_WATCH'] = df_final[watching].sum(axis=1)
df_final['SUM_TEST'] = df_final[testing].sum(axis=1)
df_final['SUM_SPEAK'] = df_final[speaking].sum(axis=1)
df_final['SUM_WRITE'] = df_final[writing].sum(axis=1)


sum_languages = ['SUM_LIST', 'SUM_READ', 'SUM_WATCH', 'SUM_TEST', 'SUM_SPEAK', 'SUM_WRITE']

df_prior = df_final.copy()

awareness = ['MEDITATION', 'BREATH']
health_five = ['MORNING', 'EYES', 'CTRAINING']
love_and_will = ['MK', 'PNW', 'ALW']
iq_dev = ['IQ', 'PROF', 'LANGUAGE']
speakingdev = ['SPEAK_DUT', 'SPEAK_ENG']
readingdev = ['SUM_READ', 'SUM_LIST']

df_prior['AWARENESS'] = df_prior[awareness].sum(axis=1)
df_prior['HEALTHFIVE'] = df_prior[health_five].sum(axis=1)
df_prior['LOVEWILL'] = df_prior[love_and_will].sum(axis=1)
df_prior['IQDEV'] = df_prior[iq_dev].sum(axis=1)
df_prior['SPEAKDEV'] = df_prior[speakingdev].sum(axis=1)
df_prior['READDEV'] = df_prior[readingdev].sum(axis=1)
df_prior['NEGATIVE'] = df_prior['FASTFOOD'] * 3 + df_prior['SWEETS'] * 2 + df_prior['COFFEE'] * 1

priors = ['DATE', 'SPORT_TIME', 'PYTHON', 'TOTAL_DUT', 'PRIOR', 'DIFFICULT', 'SPEAKDEV', 'READDEV', 'IQDEV',
          'STRENGHT_POINTS', 'MEETS', 'WRITE_J', 'AWARENESS', 'HEALTHFIVE', 'LOVEWILL', 'NEGATIVE']

df_prior = df_prior[priors]

df = df_prior.copy()

bins_dict = {
    'SPORT_TIME': [0, 15, 25, 45, 90, float('inf')],
    'PYTHON': [0, 25, 45, 90, 120, float('inf')],
    'TOTAL_DUT': [0, 15, 25, 45, 90, float('inf')],
    'PRIOR': [0, 15, 45, 120, 240, float('inf')],
    'DIFFICULT': [0, 10, 30, 60, 90, float('inf')],
    'SPEAKDEV': [0, 10, 30, 45, 60, float('inf')],
    'READDEV': [0, 10, 25, 45, 90, float('inf')],
    'IQDEV': [0, 10, 25, 45, 90, float('inf')],
    'STRENGHT_POINTS': [0, 400, 550, 700, 850, float('inf')],
    'MEETS': [0, 30, 60, 90, 120, float('inf')],
    'WRITE_J': [0, 10, 25, 45, 60, float('inf')],
    'AWARENESS': [0, 5, 8, 15, 30, float('inf')],
    'HEALTHFIVE': [0, 1.1, 2.1, 3.1, 4.1, float('inf')],
    'LOVEWILL': [0, 1.1, 2.1, 3.1, 4.1, float('inf')]}

# Define the labels for the bins
labels = [1, 2, 3, 4, 5]

# Process each column to create level columns based on the predefined bins
for col in bins_dict.keys():
    df[f'{col}_LEVEL'] = pd.cut(df[col], bins=bins_dict[col], labels=labels, include_lowest=False, right=False)

    # Convert the LEVEL column to object type to handle zero values
    df[f'{col}_LEVEL'] = df[f'{col}_LEVEL'].astype(object)

    # Handle zero values separately
    df.loc[df[col] == 0, f'{col}_LEVEL'] = 0

    # Convert the LEVEL column back to integer type
    df[f'{col}_LEVEL'] = df[f'{col}_LEVEL'].astype(int)

neg_bins_dict = [- float('inf'), -10, -5, -3, 0, 1]

# Apply pd.cut to create level column for 'NEGATIVE'
df['NEGATIVE_LEVEL'] = pd.cut(df['NEGATIVE'], bins=neg_bins_dict, labels=labels, include_lowest=False, right=True)

# Convert the NEGATIVE_LEVEL column to object type to handle specific values
df['NEGATIVE_LEVEL'] = df['NEGATIVE_LEVEL'].astype(object)

# Handle zero values separately
df.loc[df['NEGATIVE'] == 0, 'NEGATIVE_LEVEL'] = 5

# Add a new category (0) to the NEGATIVE_LEVEL column
df['NEGATIVE_LEVEL'] = df['NEGATIVE_LEVEL'].astype('category')
df['NEGATIVE_LEVEL'] = df['NEGATIVE_LEVEL'].cat.add_categories([0])

# Handle values less than -20 separately
df.loc[df['NEGATIVE'] < -20, 'NEGATIVE_LEVEL'] = 0

# Convert the NEGATIVE_LEVEL column back to integer type
df['NEGATIVE_LEVEL'] = df['NEGATIVE_LEVEL'].astype(int)

# Select 'DATE' column and the last fifteen columns
columns_to_keep = ['DATE'] + df.columns[-15:].tolist()
df_priors_final = df[columns_to_keep]

# Define the weights for each column
weights_p = {
    'SPORT_TIME_LEVEL': 2,
    'PYTHON_LEVEL': 1.5,
    'TOTAL_DUT_LEVEL': 2,
    'PRIOR_LEVEL': 1.5,
    'DIFFICULT_LEVEL': 0.5,
    'SPEAKDEV_LEVEL': 1.5,
    'READDEV_LEVEL': 0.5,
    'IQDEV_LEVEL': 1,
    'STRENGHT_POINTS_LEVEL': 0.5,
    'MEETS_LEVEL': 1,
    'WRITE_J_LEVEL': 0.5,
    'AWARENESS_LEVEL': 0.5,
    'HEALTHFIVE_LEVEL': 1,
    'LOVEWILL_LEVEL': 0.5,
    'NEGATIVE_LEVEL': 0.5
}

weights_s = {
    'SPORT_TIME_LEVEL': 1,
    'PYTHON_LEVEL': 1,
    'TOTAL_DUT_LEVEL': 1,
    'PRIOR_LEVEL': 1,
    'DIFFICULT_LEVEL': 1,
    'SPEAKDEV_LEVEL': 1,
    'READDEV_LEVEL': 1,
    'IQDEV_LEVEL': 1,
    'STRENGHT_POINTS_LEVEL': 1,
    'MEETS_LEVEL': 1,
    'WRITE_J_LEVEL': 1,
    'AWARENESS_LEVEL': 1,
    'HEALTHFIVE_LEVEL': 1,
    'LOVEWILL_LEVEL': 1,
    'NEGATIVE_LEVEL': 1
}

# Calculate the normalization factor
normalization_factor = 100 / (15 * 5)

# Apply weights to each column and calculate the score
# Calculate SCORE
df_priors_final.loc[:, 'SCORE'] = sum(
    df_priors_final[col] * weight * normalization_factor for col, weight in weights_p.items()
)

df_priors_final.loc[:, 'SCORE_S'] = sum(
    df_priors_final[col] * weight * normalization_factor for col, weight in weights_s.items()
)



df_p_m = df_priors_final.copy()

df_p_m['DATE'] = pd.to_datetime(df_p_m['DATE'], format='%d.%m.%Y')
df_p_m['month_year'] = df_p_m['DATE'].dt.to_period('M')
df_p_m = df_p_m.pivot_table(index='month_year', aggfunc='mean')

# weeks results

df_p_w = df_priors_final.copy()

# Ensure the 'date' column is in datetime format
df_p_w['DATE'] = pd.to_datetime(df_p_w['DATE'],format='%d.%m.%Y')

# Set the 'date' column as the index
df_p_w.set_index('DATE', inplace=True)

# Resample data by week and aggregate
weekly_data_p = df_p_w.resample('W').mean()

# month results
df_month = df_final.copy()

df_month['DATE'] = pd.to_datetime(df_month['DATE'],format='%d.%m.%Y')
df_month['month_year'] = df_month['DATE'].dt.to_period('M')

df_month.drop(columns=['DATE'], inplace=True)


writer = pd.ExcelWriter('data_files/df_month.xlsx', engine='xlsxwriter')
df_month.to_excel(writer, sheet_name='df_month')
writer.close()

#######

#df_month_pivot = df_month.pivot_table(index='month_year', aggfunc='sum')




# weeks results

df_weeks = df_final.copy()

# Ensure the 'date' column is in datetime format
df_weeks['DATE'] = pd.to_datetime(df_weeks['DATE'],format='%d.%m.%Y')

# Set the 'date' column as the index
df_weeks.set_index('DATE', inplace=True)

# Resample data by week and aggregate
#weekly_data = df_weeks.resample('W').sum()

#2023 without the partial 1st week
#weekly_data = weekly_data.iloc[1:]
#max_values = weekly_data.max()

#2024 only
#weekly_data_2024 = weekly_data.iloc[52:]
#max_values_2024 = weekly_data_2024.max()

#last_row = weekly_data.iloc[-1]


new_column_order = separate_timing + imposed_timing + additional +  yes_no + times + negative + quantities + distance + sport_additional + typ_lang_eng + typ_lang_rus + typ_lang_spa + typ_lang_dut + sum_languages

new_column_order_final = ['DATE'] + new_column_order + ['START', 'FINISH']

df_final = df_final[new_column_order_final]
#df_month_pivot = df_month_pivot[new_column_order]
#weekly_data = weekly_data[new_column_order]


last_row = df.iloc[-1]
split_index = (len(last_row) - 1) // 2
date_column = last_row.index[0]
columns_part1 = last_row.index[1:split_index + 1]
columns_part2 = last_row.index[split_index + 1:]
print(f"{date_column}: {last_row[date_column]}")
max_length = max(len(columns_part1), len(columns_part2))
for i in range(max_length):
    col1_text = f"{columns_part1[i]}: {last_row[columns_part1[i]]}" if i < len(columns_part1) else ""
    col2_text = f"{columns_part2[i]}: {last_row[columns_part2[i]]}" if i < len(columns_part2) else ""
    print(f"{col1_text:<30} ------------- {col2_text}")


last_row_p = df_priors_final.iloc[-1]

value = float(last_row_p['SCORE'])
rounded_value = round(value, 2)

print('\nDAY SCORE = ', rounded_value)

#Sort the DataFrame by 'DATE' from newest to oldest

# Convert 'DATE' column to datetime with the specific format
df_final['DATE'] = pd.to_datetime(df_final['DATE'], format='%d.%m.%Y')
df_priors_final['DATE'] = pd.to_datetime(df_priors_final['DATE'], format='%d.%m.%Y')
df['DATE'] = pd.to_datetime(df['DATE'], format='%d.%m.%Y')

df_final = df_final.sort_values(by='DATE', ascending=False)
df_priors_final = df_priors_final.sort_values(by='DATE', ascending=False)
#df_month_pivot = df_month_pivot.sort_values(by='month_year', ascending=False)
#weekly_data= weekly_data.sort_values(by='DATE', ascending=False)
df_p_m  = df_p_m.sort_values(by='month_year', ascending=False)
weekly_data_p= weekly_data_p.sort_values(by='DATE', ascending=False)
df = df.sort_values(by='DATE', ascending=False)

writer = pd.ExcelWriter('data_files/days_stat.xlsx')
df_final.to_excel(writer, sheet_name='days', index=False)
#df_month_pivot.to_excel(writer, sheet_name='months')
#weekly_data.to_excel(writer, sheet_name='weeks')
df_priors_final.to_excel(writer, sheet_name='priors', index=False)
df_p_m.to_excel(writer, sheet_name='priors_months')
weekly_data_p.to_excel(writer, sheet_name='priors_weeks')
df.to_excel(writer, sheet_name='df', index=False)

art_final.to_excel(writer, sheet_name='art', index=False)
writer.close()




