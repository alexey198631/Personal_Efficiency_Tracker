import gspread
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from data_files import config

filename = config.SERVICE_ACCOUNT
dairy_file = config.DAYS
gbook = config.GBOOK
glist = config.GSHEET
elist = config.ELIST

# Prepare data frame for Excel workbook sheet
def df_from_sheet(wkb, sht_name):
    # Select the sheet you want to modify
    sheet = wkb[sht_name]

    # Save data frame from sheet data
    data_days = sheet.values
    cols = next(data_days)[0:]
    df = pd.DataFrame(data_days, columns=cols)
    return df


# Write the updated DataFrame to the worksheet with the same format
def same_format_sheet(wkb, sht, df):
    # Create a new worksheet with the same formatting as the existing worksheet
    new_worksheet = wkb.create_sheet('Temp')
    new_worksheet.sheet_format = sht.sheet_format
    new_worksheet.sheet_properties = sht.sheet_properties
    new_worksheet.page_setup = sht.page_setup

    # Write the updated DataFrame to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        new_worksheet.append(r)

    # Delete all existing rows in the worksheet
    sht.delete_rows(1, sht.max_row)
    for row in new_worksheet.iter_rows():
        sht.append([cell.value for cell in row])

    # Delete the temporary worksheet
    wkb.remove(new_worksheet)

    # Save the changes to the Excel file
    wkb.save(dairy_file)

    # Close the workbook
    wkb.close()


# Open a Google sheet
gc = gspread.service_account(filename)
spr = gc.open(gbook).worksheet(glist)

# Get all the data from the sheet as a list of lists
data = spr.get_all_values()

# Convert the data to a pandas dataframe
df_google = pd.DataFrame(data[1:], columns=data[0])

# Search for the last not empty element in DAY column
last_index = df_google['DAY'].tolist().index('') - 1
df_google = df_google.loc[:last_index, :]

# Load the Days Diary xlsx workbook
workbook = openpyxl.load_workbook(dairy_file)

# Create Data Frame from Excel file
df_days = df_from_sheet(workbook, elist)
# Save columns to a list
cols = list(df_days.columns)

# Rename all columns which have different name because in Google Sheet file modified names are being used
col_names_correct = {}
for i in range(len(cols)):
    if cols[i] != data[0][i]:
        col_names_correct[data[0][i]] = cols[i]
        df_google = df_google.rename(columns=col_names_correct)

# Loop over each column and convert it to the desired data type
for col in df_google.columns:
    # Check if the data type of the column is a string ('object')
    col_dtype = df_days[col].dtype
    # Convert the column to dtype of original dataframe
    if col_dtype == 'datetime64[ns]':
        df_google[col] = pd.to_datetime(df_google[col], format='%d/%m/%Y')
    elif col_dtype == 'float':
        # Apply the transformation to the entire column
        df_google[col] = df_google[col].apply(lambda x: float(x.replace(',', '.')))
        df_google[col] = df_google[col].astype(col_dtype)
    else:
        df_google[col] = df_google[col].astype(col_dtype)

# Write the updated DataFrame to the worksheet with the same format
same_format_sheet(workbook, workbook[elist], df_google)

# Prepare reference Sheet for Phase 2
weeks_df = df_from_sheet(workbook, 'week')

# Filter rows which exists in df from Google
weeks_df['InBoth'] = weeks_df['DATE'].isin(df_google['DATE']).astype(int)

# Fill in columns in df for dates that are present in both dataframes
weeks_df.loc[weeks_df['InBoth'] == 1, 'WEEK#'] = weeks_df.loc[weeks_df['InBoth'] == 1, 'WEEK#T']
weeks_df.loc[weeks_df['InBoth'] == 1, 'MONTH#'] = weeks_df.loc[weeks_df['InBoth'] == 1, 'MONTH#T']

# Drop the 'InBoth' column
weeks_df = weeks_df.drop('InBoth', axis=1)

# Write the updated DataFrame to the worksheet with the same format
same_format_sheet(workbook, workbook['week'], weeks_df)

gtiming = config.TIMINGBOOK
gart = config.ARTBOOK
gday = config.DAYBOOK
gweek = config.WEEKBOOK
gmonth = config.MONTHBOOK

def sheet_to_df(sheet_var):
    # Open a Google sheet
    spr = gc.open(gbook).worksheet(sheet_var)
    # Get all the data from the sheet as a list of lists
    data = spr.get_all_values()
    # Convert the data to a pandas dataframe
    df_google = pd.DataFrame(data[1:], columns=data[0])
    return df_google




days = sheet_to_df(gday)
weeks = sheet_to_df(gweek)
months = sheet_to_df(gmonth)
art = sheet_to_df(gart)

# Define an integer format string
number_format = '#,##0.00'

writer = pd.ExcelWriter('data_files/month.xlsx', engine='xlsxwriter')
days.to_excel(writer, sheet_name='Sheet1', index=False, float_format=number_format)
weeks.to_excel(writer, sheet_name='Sheet2', index=False, float_format=number_format)
months.to_excel(writer, sheet_name='Sheet3', index=False, float_format=number_format)
art.to_excel(writer, sheet_name='Sheet4', index=False, float_format=number_format)
writer.save()